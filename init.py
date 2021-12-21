import os # name
from openpyxl import Workbook, load_workbook


def obtener_comando_borrar():
	''' Chequeamos si el SO es Windows o Linux para ver que comando utilizaremos con shutil. '''
	if os.name == "nt": # == Era Windows
		return "cls" # <--- Windows
	else:
		return "clear" # <--- Linux


def abrir_excel():
	try:
		wb = load_workbook("Compras confirmadas.xlsx")
		ws = wb.active
	except FileNotFoundError:
		wb = Workbook()
		ws = wb.active
		ws.append(["Cliente", "Fecha", "Combo S", "Combo D", "Combo T", "Flurby", "Total"])

	return wb, ws